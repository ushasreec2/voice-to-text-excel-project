from flask import Flask, request, jsonify, send_from_directory, send_file
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from werkzeug.utils import secure_filename
import os
import re
from flask_cors import CORS

app = Flask(__name__)
CORS(app)
UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def parse_command(text):
    text = text.lower().strip()

    # Pattern 1: insert value in cell D2
    match = re.match(r"insert (.+) in cell ([a-z]+)(\d+)", text)
    if match:
        raw_value = match.group(1).strip()
        col_letter = match.group(2).upper()
        row = int(match.group(3))
        col = column_index_from_string(col_letter)

        # Check if row or column exceeds Excel's max limit
        if col > 16384 or row > 1048576:
            return {"error": "Cell location exceeds Excel's supported range."}

        if raw_value.startswith("equal"):
            raw_value = raw_value.replace("equal", "").strip()

            def handle_range_formula(keyword, func_name):
                parts = raw_value.replace(keyword, "").strip().split(" to ")
                if len(parts) == 2:
                    start = parts[0].strip().upper()
                    end = parts[1].strip().upper()
                    return f"={func_name}({start}:{end})"
                return None

            if raw_value.startswith("sum"):
                value = handle_range_formula("sum", "SUM")
            elif raw_value.startswith("average"):
                value = handle_range_formula("average", "AVERAGE")
            elif raw_value.startswith("min"):
                value = handle_range_formula("min", "MIN")
            elif raw_value.startswith("max"):
                value = handle_range_formula("max", "MAX")
            elif raw_value.startswith("count"):
                value = handle_range_formula("count", "COUNT")
            elif raw_value.startswith("standard deviation"):
                value = handle_range_formula("standard deviation", "STDEV")
            elif raw_value.startswith("round"):
                match = re.match(r"round ([a-z]+\d+|\d+(\.\d+)?) to (\d+) decimal", raw_value)
                if match:
                    value = match.group(1).upper()
                    decimals = match.group(3)
                    value = f"=ROUND({value}, {decimals})"
            elif raw_value.startswith("ceil") or raw_value.startswith("ceiling"):
                match = re.match(r"(ceil|ceiling) ([a-z]+\d+|\d+(\.\d+)?)", raw_value)
                if match:
                    val = match.group(2).upper()
                    value = f"=ROUNDUP({val}, 0)"
            elif raw_value.startswith("floor"):
                match = re.match(r"floor ([a-z]+\d+|\d+(\.\d+)?)", raw_value)
                if match:
                    val = match.group(1).upper()
                    value = f"=ROUNDDOWN({val}, 0)"
            elif raw_value.startswith("power"):
                match = re.match(r"power ([a-z]+\d+|\d+(\.\d+)?) to ([a-z]+\d+|\d+(\.\d+)?)", raw_value)
                if match:
                    base = match.group(1).upper()
                    exponent = match.group(3).upper()
                    value = f"=POWER({base}, {exponent})"
            elif raw_value.startswith("sqrt") or raw_value.startswith("square root"):
                match = re.match(r"(sqrt|square root) of ([a-z]+\d+|\d+(\.\d+)?)", raw_value)
                if match:
                    number = match.group(2).upper()
                    value = f"=SQRT({number})"
            else:
                value = "=" + raw_value.replace("plus", "+").replace("minus", "-").replace("times", "*").replace("divided by", "/")
        else:
            value = raw_value

        return {
            "value": value,
            "row": row,
            "column": col
        }

    # Pattern 2: insert value in row 1 column 2
    match = re.match(r"insert (.+) in row (\d+) column (\d+)", text)
    if match:
        value = match.group(1).strip()
        row = int(match.group(2))
        col = int(match.group(3))

        # Check if row exceeds Excel's max limit
        if row > 1048576:
            return {"error": "Row number exceeds Excel's maximum limit of 1,048,576."}

        return {
            "value": value,
            "row": row,
            "column": col
        }

    return None

@app.route('/')
def home():
    return send_file('index.html')
@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files and 'filename' not in request.form:
        return jsonify({'error': 'Missing file or filename'}), 400
    
    if 'voice_text' not in request.form:
        return jsonify({'error': 'Missing voice input'}), 400

    voice_input = request.form['voice_text']
    parsed = parse_command(voice_input)

    if not parsed or "error" in parsed:
        return jsonify({'error': parsed.get("error", "Could not parse command")}), 400

    # File handling
    if 'file' in request.files:
        file = request.files['file']
        filename = secure_filename(file.filename)
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        file.save(filepath)
    else:
        filename = secure_filename(request.form['filename'])
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        if not os.path.exists(filepath):
            return jsonify({'error': 'File not found in server'}), 400

    wb = load_workbook(filepath)
    ws = wb.active

    # 👉 Insert Normal Value / Formula
    if "value" in parsed:
        ws.cell(row=parsed["row"], column=parsed["column"], value=parsed["value"])
        wb.save(filepath)

    # 👉 Generate Sheet Preview (first 10x10 cells)
    preview_data = []
    for r in range(1, 11):
        row_data = []
        for c in range(1, 11):
            row_data.append(ws.cell(row=r, column=c).value)
        row_data = [cell if cell is not None else "" for cell in row_data]
        preview_data.append(row_data)

    return jsonify({
        'message': f"Inserted value in {get_column_letter(parsed['column'])}{parsed['row']}",
        'value': parsed["value"],
        'filename': filename,
        'sheet_preview': preview_data  # 👈 now frontend gets the expected data
    }), 200



@app.route('/download/<filename>')
def download_file(filename):
    filename = secure_filename(filename)  # Ensure it's a safe path
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    if os.path.exists(filepath):
        return send_from_directory(UPLOAD_FOLDER, filename, as_attachment=True)
    else:
        return jsonify({'error': 'File not found'}), 404


if __name__ == '__main__':
    app.run(debug=True)
